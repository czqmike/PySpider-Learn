#-*- coding:utf-8 -*-

import requests
from pyquery import PyQuery as pq
from openpyxl import Workbook
import openpyxl.styles as sty
import datetime
import os

class gach:
    __time__ = ""
    __title__ = ""
    __text__ = ""

# return: list of ids, get from web
def getID():
    print("正在获取id列表...")
    url = "http://www.mihoyo.com/news/getNotice"
    result = requests.get(url)
    if (result.status_code == 200):
        json = result.json()
        items = json.get('data').get('gach')
        id = []
        for item in items:
            id.append(item.get('id'))
    else:
        print("获取ID列表失败！")
    return id

# para:list of ids(str)
# return: a list of gach
def getGachByID(id):
    ga = []
    print("正在获取祈愿信息...")
    i = 1
    for item in id:
        url = "http://www.mihoyo.com/news/getNoticeByID?id=" + item
        result = requests.get(url)
        if (result.status_code == 200):
            data = result.json().get('data')
            # defending error of type
            if data and data.get('type') == '2':
                print("正在获取第%d条祈愿信息..." % i)
                i += 1
                g = gach()
                g.time = data.get('time')
                g.title = data.get('title')

                # use PyQuery select the names of every gach information
                html = data.get('text')
                doc = pq(html)
                items = doc('tr td')
                g.text = items.text()

                ga.append(g)
        else:
            print("获取祈愿信息失败！")
    return ga

def str_to_datetime(str,format='%Y-%m-%d %H:%M:%S'):
    return datetime.datetime.strptime(str,format)

def datetime_to_str(dt=datetime.datetime.now()):
    return dt.date().strftime('%Y_%m_%d')

junk = [" 崩坏世界的歌姬", " 禁断学园", " 终焉之种", " 源初虹石", " 芙蕾雅之梦", " 负罪之羽", "喰时之种", ">>>>",
        " 魔神结晶", "集齐50颗可合成", "集齐30颗可合成", " 遗迹之心", " 量子核心", "CD：崩坏世界的歌姬", "CD:崩坏世界的歌姬"]

print("本程序可自动获取祈愿信息并保存在当前文件夹下，可以获取最近4个月或是服务器上所有的记录")
print("--若要获取最近4个月的记录，请直接按Enter键，获取时间大约1分钟")
print("--若要获取所有记录，请输入 woquandouyao 后按 Enter，获取时间大约30分钟！！！（慎重）")
choice = input("请输入选项：")

# fill id to id_list
if choice == "woquandouyao":
    # fill id_list from 1 to nearest_id
    i = int(getID()[0])
    id_list = []
    while (i >= 1):
        id_list.append(str(i))
        i -= 1
else:
    id_list = getID()

#get gach information by IDs in list
gach_list = getGachByID(id_list)
gach_list  = list(gach_list)
for item in gach_list:      # delete string we do not need
    for ju in junk:
        item.text = item.text.replace(ju, "")
    # print(item.time, item.title, item.text)

# save data to a excel(xlsx)
print("正在填充excel表...")
wb = Workbook()
Sheet = wb.active
Sheet.cell(1, 1).value = "日期"
Sheet.cell(1, 2).value = "标题"
Sheet.cell(1, 3).value = "蛋池内容"
for i in range(2, len(gach_list) + 2):
    Sheet.cell(i, 1).value = gach_list[i - 2].time
    Sheet.cell(i, 2).value = gach_list[i - 2].title

    # set colors of cells by the type of gach
    s = (str)(gach_list[i - 2].title)
    if s.find("公主") != -1:
        Sheet.cell(i, 2).fill=sty.PatternFill(fill_type='solid',fgColor="ED7D31")
    elif s.find("使魔") != -1:
        Sheet.cell(i, 2).fill=sty.PatternFill(fill_type='solid',fgColor="5B9BD5")
    elif s.find("魔女") != -1:
        Sheet.cell(i, 2).fill=sty.PatternFill(fill_type='solid',fgColor="7030A0")
    else:
        Sheet.cell(i, 2).fill=sty.PatternFill(fill_type='solid',fgColor="FF6699")

    text = gach_list[i - 2].text
    text_list = text.split()
    for j in range(3, len(text_list) + 3):
        Sheet.cell(i, j).value = text_list[j - 3]

# save excel
print("正在保存excel表...")
start_time = datetime_to_str(str_to_datetime(gach_list[len(gach_list) - 1].time))
end_time = datetime_to_str(str_to_datetime(gach_list[0].time))
if choice != "woquandouyao":
    save_name = "神器up周期表 " + start_time + '~' + end_time + ".xlsx"
else:
    save_name = "神器up周期表总 " + 'start_time~' + end_time + ".xlsx"
wb.save(save_name)
print("保存成功！")
os.system("pause")


